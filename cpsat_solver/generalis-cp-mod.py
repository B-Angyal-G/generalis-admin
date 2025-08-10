from ortools.sat.python import cp_model
import copy as c
import os
from openpyxl import *


def request_check(ORIG_TABLE):
    DAYS = len(ORIG_TABLE[0]) - 6
    ERROR = 0

    ### ### ELLENORZES, HOGY PONTOSAN ANNYI KIOSZTANDO MUSZAK VAN-E AHANY NAP
    sum_shifts = sum( row[5] for row in ORIG_TABLE )
    
    ### MAX, MIN NAPPALOS ES EJSZAKAS MUSZAKOK SZAMAINAK ELLENORZESE
    sum_day_shift_min = sum( row[1] for row in ORIG_TABLE )
    sum_day_shift_max = sum( row[2] for row in ORIG_TABLE )
    sum_night_shift_min = sum( row[3] for row in ORIG_TABLE )
    sum_night_shift_max = sum( row[4] for row in ORIG_TABLE )

    if sum_day_shift_min > 2 * DAYS:
        print('Nappalos műszakokból legalább beírandó több, mint amennyit a hónapban ki kell osztani!')
        ERROR = 1
    if sum_day_shift_max < 2 * DAYS:
        print('Nappalos műszakokból legfeljebb beírandó kevesebb, mint amennyit a hónapban ki kell osztani!')
        ERROR = 1

    if sum_night_shift_min > DAYS:
        print('Éjszakás műszakokból legalább beírandó több, mint amennyit a hónapban ki kell osztani!')
        ERROR = 1
    if sum_night_shift_max < DAYS:
        print('Éjszakás műszakokból legfeljebb beírandó kevesebb, mint amennyit a hónapban ki kell osztani!')
        ERROR = 1

    if sum_shifts != 3 * DAYS:
        print('Nem megfelelő az összes műszakszám kiosztása!')
        ERROR = 1

    ### ### TABLAZAT ELLENORZESE
    for row in ORIG_TABLE:
        day_shift = 0
        night_shift = 0
        possible_days = 0

        ### TABLAZAT KITOLTESE LOGIKAILAG NINCS-E ELRONTVA
        if row[1] > row[2]:
            print(row[0], 'adminnak több minimum nappalos műszakot kell megadni, mint maximumot!')
            ERROR = 1
        if row[3] > row[4]:
            print(row[0], 'adminnak több minimum éjszakás műszakot kell megadni, mint maximumot!')
            ERROR = 1

        for day in range(6, len(row)):
            ### ### ELLENORZES, HOGY NEM ADTAK-E EGY EMBERNEK TOBB FIX NAPOT
            ### ### NAPPALRA VAGY EJSZAKARA, MINT AMENNYIT LEHETNE
            ### ### ES
            ### ### ELLENORZES, HOGY NEM ADTAK-E VALAKINEK
            ### ### EJSZAKA UTAN NAPPALT
                if row[day] in {'n', 'N'}:
                    day_shift += 1
                    if row[day - 1] in {'e', 'E', 'é', 'É'}:
                        print(row[0], day - 1, 'napokon éjszakás után nappalos!')
                        ERROR = 1

                if row[day] in {'e', 'E', 'é', 'É'}:
                    night_shift += 1
                
                if row[day] == None:
                    possible_days += 1

        ### HIBAUZENETEK
        if day_shift > row[2]:
            print(row[0], 'adminnak több fix nappalos műszak lett beírva, mint amennyi maximum lehet!')
            ERROR = 1

        if night_shift > row[4]:
            print(row[0], 'adminnak több fix éjszakás műszak lett beírva, mint amennyi maximum lehet!')
            ERROR = 1

        if possible_days < row[5]:
            print(row[0], 'adminnak nincs annyi beosztható napja, ahány napra be kéne osztani!')
            ERROR = 1
                

    ### ### ELLENORZES, HOGY NEM OSZTOTTAK-E BE TOBB EMBERT EGY NAPON
    ### ### NAPPALRA VAGY EJSZAKARA, MINT AMENNYIT LEHETNE
    for day in range(6, len(ORIG_TABLE[0])):
        s_day = 0
        s_night = 0
        for admin in range(len(ORIG_TABLE)):
            if ORIG_TABLE[admin][day] in {'n', 'N'}:
                s_day += 1
            if ORIG_TABLE[admin][day] in {'e', 'E', 'é', 'É'}:
                s_night += 1
        if s_day > 2:
            print(day - 5, 'napon több fix nappalos műszak lett beírva, mint 2!')
            ERROR = 1
        if s_night > 1:
            print(day - 5, 'napon több fix éjszakás műszak lett beírva!')
            ERROR = 1

    if ERROR:
        return 0
    return 1


class SolutionPrinter(cp_model.CpSolverSolutionCallback):
    ### MEGOLDAS VISSZAHIVO OSZTALY, AMELY KIIRJA A TALALT MEGOLDASOKAT
    out_workbook = Workbook()
    out_sheet = out_workbook.active
    schedules = out_sheet.title
    act_row = -1

    def __init__(self, admins_name, shifts, all_admins, num_days, all_shifts, sat_orig, sun_orig, solution_limit=None):
        cp_model.CpSolverSolutionCallback.__init__(self)
        self._admins_name = admins_name
        self._shifts = shifts
        self._all_admins = all_admins
        self._num_days = num_days
        self._all_shifts = all_shifts
        self._sat_orig = sat_orig
        self._sun_orig = sun_orig
        self._solution_count = 0
        self._solution_limit = solution_limit
        self.all_found_schedules = [] # ITT TAROLJUK AZ OSSZES TALALT MEGOLDAST

    def OnSolutionCallback(self):
        ### EZT A METODUST HIVJA MEG A SOLVER MINDEN TALALT MEGOLDASKOR
        self._solution_count += 1
        print(f"\n--- Megoldás {self._solution_count} ---")
        current_schedule = []

        SolutionPrinter.act_row += 2
        table_fin = list()
        for a in self._all_admins:
            admins_shifts_today = [self._admins_name[a]]
            for d in range(self._num_days):
                OK = 0
                for s in self._all_shifts:
                    if self.Value(self._shifts[(a, d, s)]) == 1:
                        if s == 0:
                            admins_shifts_today.append('N')
                            OK = 1
                        elif s == 1:
                            admins_shifts_today.append('E')
                            OK= 1
                if not OK:
                    admins_shifts_today.append(' ')
            act_col = 1
            for i in admins_shifts_today:
                cell = SolutionPrinter.out_sheet.cell(row = SolutionPrinter.act_row, column = act_col)
                cell.value = i
                act_col += 1
            SolutionPrinter.act_row += 1
            table_fin.append(admins_shifts_today)
            print(admins_shifts_today)
        rating = eval_table(table_fin, self._sat_orig, self._sun_orig)
        cell = SolutionPrinter.out_sheet.cell(row = SolutionPrinter.act_row, column = 1)
        cell.value = 'Értékelés:'
        cell = SolutionPrinter.out_sheet.cell(row = SolutionPrinter.act_row, column = 2)
        cell.value = rating
        print(rating)
        print()
        SolutionPrinter.out_workbook.save('./admin_schedule_tmp.xlsx')

        self.all_found_schedules.append(current_schedule)

        # HA BE VAN ALLITVA MEGOLDAS LIMIT, ES ELERTUK, AKKOR ALLITSUK LE A KERESEST
        if self._solution_limit is not None and self._solution_count >= self._solution_limit:
            print(f"Elértük a {self._solution_limit} megoldás limitet. Leállítás.")
            self.StopSearch()

    def solution_count(self):
        return self._solution_count


def read_requests(ORIG_TABLE, all_admins, all_days, all_shifts):
    shift_requests = [[[1 for s in all_shifts] for d in all_days] for a in all_admins]

    all_exception = {'x', 'X', 'y', 'Y'}
    day_exception_current = {'xn', 'xN', 'XN', 'Xn', 'nx', 'nX', 'Nx', 'NX', 'e', 'E', 'é', 'É'}
    day_current = {'n', 'N'}

    night_exception_current = {'xe', 'xE', 'XE', 'Xe', 'ex', 'eX', 'Ex', 'EX', 'n', 'N',
                               'xé', 'xÉ', 'XÉ', 'Xé', 'éx', 'éX', 'Éx', 'ÉX'}
    night_exception_next = {'x', 'X', 'xn', 'xN', 'XN', 'Xn', 'nx', 'nX', 'Nx', 'NX', 'n', 'N'}
    night_current = {'e', 'E', 'é', 'É'}

    for a in all_admins:
        for d in all_days:
            if ORIG_TABLE[a][d] in day_current:
                shift_requests[a][d][0] = 2
                shift_requests[a][d][1] = 0
            elif ORIG_TABLE[a][d] in night_current:
                shift_requests[a][d][0] = 0
                shift_requests[a][d][1] = 2
            elif ORIG_TABLE[a][d] in all_exception:
                shift_requests[a][d][0] = 0
                shift_requests[a][d][1] = 0
            elif ORIG_TABLE[a][d] in day_exception_current:
                shift_requests[a][d][0] = 0
            elif ORIG_TABLE[a][d] in night_exception_current:
                shift_requests[a][d][1] = 0

            if d < len(all_days) - 1:
                if ORIG_TABLE[a][d] in night_current:
                    shift_requests[a][d + 1][0] = 0
            if d > 0:
                if ORIG_TABLE[a][d] in night_exception_next:
                    shift_requests[a][d - 1][1] = 0

    return shift_requests


def pattern_count(row, pattern):
    row_tmp = c.deepcopy(row)
    row_tmp = row_tmp.replace('N', '*')
    row_tmp = row_tmp.replace('E', '*')
    l = len(pattern)

    s = 0
    current = ''
    for i in range(l):
        current += row_tmp[1 + i]
    if current == pattern:
        s += 1

    for i in range(1 + l, len(row_tmp)):
        current = current[1:]
        current += row_tmp[i]
        if current == pattern:
            s += 1
    return s


def eval_table(table_fin, sat_orig, sun_orig):
    PRINT = 0

    VALUE = 10000
    pattern = [['.**.', 35], ['.***.', 170], ['.**.*.', 40], ['.**.**.', 80], ['.***..**.', 115]]

    ### MINDEN EGYES ADMINHOZ TARTOZIK EGY KETELEMU LISTA,
    ### MELYNEK AZ ELSO ELEME AZ OSSZES MUSZAKJA, A MASODIK A HETVEGERE ESO MUSZAKJAI
    admin_shifts = [ [0, 0] for i in range(len(table_fin)) ]

    table_string = ['.' for i in range(len(table_fin))]
    for admin in range(len(table_fin)):
        s = 0
        w = 0
        for day in range(1, len(table_fin[0])):
            if table_fin[admin][day] == ' ':
                table_string[admin] += '.'
            elif table_fin[admin][day] in {'N', 'E'}:
                s += 1
                table_string[admin] += table_fin[admin][day]
                if day in sat_orig or day in sun_orig:
                    w += 1
        admin_shifts[admin][0] = s
        admin_shifts[admin][1] = w
        table_string[admin] += '.'

    weekend = 0
    weekend += len(sat_orig)
    weekend += len(sun_orig)
    weekend *= 3
    sum_shifts = 3 * (len(table_fin[0]) - 1)
    for admin in admin_shifts:
        opt_weekend = admin[0]*weekend/sum_shifts
        w = abs(admin[1] - opt_weekend)
        if w > 2:
            VALUE -= 120
        elif w > 1:
            VALUE -= 50
        elif w > 0:
            VALUE -= 20

    for row in table_string:
        for p in pattern:
            VALUE -= pattern_count(row, p[0])*p[1]

    if PRINT:
        for i in table_fin:
            print(i)
        print()
        print(VALUE)
        print()
        for i in admin_shifts:
            print(i)
        print()
        print('-'*100)
        print()

    return VALUE


def main() -> None:
    ### XLSX IMPORTALASA
    workbook = load_workbook(filename="./requests_admin_cp.xlsx", data_only = True)

    # ELSO SHEET
    worksheet = workbook.worksheets[0]

    # 2D-S LISTAVA KONVERTALAS
    row_list = []
    for r in worksheet.rows:
        column = [cell.value for cell in r]
        row_list.append(column)

    # NAPOK SZAMA
    num_days = row_list[0][1]
    solution_limit = row_list[2][1]

    # BEOLVASOTT TABLAZAT
    # num_days + 6: OSZLOPOK SZAMA A KERESEK TABLAZAT ELOTT
    # row in row_list[4:len(row_list)]]-ban 4: TABLAZAT ELOTTI SOROK SZAMA
    ORIG_TABLE = [row[0:num_days + 6] for row in row_list[5:len(row_list)]]

    # ADMINOK
    admins_name = [row[0] for row in ORIG_TABLE]
    
    ### MUSZAKSZAMOK LISTAJA
    day_shifts_min = [row[1] for row in ORIG_TABLE]
    day_shifts_max = [row[2] for row in ORIG_TABLE]
    night_shifts_min = [row[3] for row in ORIG_TABLE]
    night_shifts_max = [row[4] for row in ORIG_TABLE]
    shifts_sum = [row[5] for row in ORIG_TABLE]

    # HETVEGEK DATUMAINAK GENERALESA KEZD.
    SATURDAY = row_list[1][1]
    sat = []
    sun = []
    if SATURDAY == 7:
        sun.append(1)
    while SATURDAY <= num_days:
        sat.append(SATURDAY)
        if SATURDAY != num_days:
            sun.append(SATURDAY + 1)
        SATURDAY += 7

    # HETVEGEK EREDETI DATUMAHOZ!!! HALMAZOK LETREHOZASA
    sat_orig = set(sat)
    sun_orig = set(sun)

    # TIME (DATUMOK) LISTABAN INDEXELES!!! MIATT 1-GYEL CSOKKENTES
    for i in range(len(sat)):
        sat[i] -= 1
    for i in range(len(sun)):
        sun[i] -= 1
    # HETVEGEK DATUMAINAK GENERALASA VEGE

    # KULSOSOK MUSZAKJAI
    s = 0
    for d in range(6, len(ORIG_TABLE[-1])):
        if ORIG_TABLE[-1][d] == None:
            if d - 6 not in sat and d - 6 not in sun:
                ORIG_TABLE[-1][d] = 'y'

    for i in ORIG_TABLE:
        print(i)

    ### ELLENORZESEK
    ### TODO : CHECK!
    if request_check(ORIG_TABLE) == 0:
        return 0
    else:
        print('Check OK!')


    # CP-SAT SOLVER
    num_admins = len(admins_name)
    num_shifts = 2
    all_admins = range(num_admins)
    all_shifts = range(num_shifts)
    all_days = range(num_days)
    shift_requests = read_requests([row[6:] for row in ORIG_TABLE], all_admins, all_days, all_shifts)

    # MODEL LETREHOZASA
    model = cp_model.CpModel()

    # MUSZAK VALTOZOK LETREHOZASA
    # shifts[(a, d, s)]: a ADMIN, d NAPON, s MUSZAKBAN DOLGOZIK-E
    shifts = {}
    for a in all_admins:
        for d in all_days:
            for s in all_shifts:
                shifts[(a, d, s)] = model.NewBoolVar(f"shift_a{a}_d{d}_s{s}")

    # MEGKOTESEK
    # MINDEN NAP PONTOSAN 2 NAPPALOS MUSZAK ES 1 EJSZAKAS MUSZAKBAN DOLGOZO ADMIN VAN
    for d in all_days:
        model.Add(sum([shifts[(a, d, 0)] for a in all_admins]) == 2)

    for d in all_days:
        model.AddExactlyOne([shifts[(a, d, 1)] for a in all_admins])


    # MINDEN ADMIN EGY NAP EGY MUSZAKBAN DOLGOZHAT
    for a in all_admins:
        for d in all_days:
            model.AddAtMostOne([shifts[(a, d, s)] for s in all_shifts])

    # EJSZAKAS UTAN NEM LEHET NAPPAL DOLGOZNI
    for a in all_admins:
        for d in range(num_days - 1):
            model.Add(shifts[(a, d + 1, 0)] == 0).OnlyEnforceIf(shifts[(a, d, 1)])

    # NEM LEHET ZSINORBAN 3-SZOR UGYANABBAN A MUSZAKBAN DOLGOZNI
    for a in all_admins:
        for d in range(num_days - 2):
            model.Add(shifts[(a, d + 2, 0)] == 0).OnlyEnforceIf([shifts[(a, d, 0)], shifts[(a, d + 1, 0)]])
            model.Add(shifts[(a, d + 2, 1)] == 0).OnlyEnforceIf([shifts[(a, d, 1)], shifts[(a, d + 1, 1)]])

    # NEM LEHET 4 MUSZAKOT FOLYAMATOSAN DOLGOZNI (ELOZO SZABALY MIATT EGYEDUL NNEE MODON VALOSULHATNA MEG)
    for a in all_admins:
        for d in range(num_days - 3):
            model.Add(shifts[(a, d + 3, 1)] == 0).OnlyEnforceIf([shifts[(a, d, 0)], shifts[(a, d + 1, 0)], shifts[(a, d + 2, 1)]])

    # TABLAZATBAN SZEREPELNEK A NAPPALOS ES EJSZAKAS MUSZAKOK MINIMUM ES MAXIMUM ERTEKEI
    for a in all_admins:
        model.Add(sum( [shifts[(a, d, 0)] for d in all_days] ) >= day_shifts_min[a])
        model.Add(sum( [shifts[(a, d, 0)] for d in all_days] ) <= day_shifts_max[a])
        model.Add(sum( [shifts[(a, d, 1)] for d in all_days] ) >= night_shifts_min[a])
        model.Add(sum( [shifts[(a, d, 1)] for d in all_days] ) <= night_shifts_max[a])
        model.Add(sum( [shifts[(a, d, s)] for d in all_days for s in all_shifts] ) == shifts_sum[a])

    
    # KERESEK A TABLAZATBOL
    for a in all_admins:
        for d in all_days:
            for s in all_shifts:
                if shift_requests[a][d][s] < 2:
                    model.Add(shifts[(a, d, s)] <= shift_requests[a][d][s])
                else:
                    model.Add(shifts[(a, d, s)] == 1)

    # EREDMENYNEK A TABLZAT
    fin_table = [[i for _ in range(num_days + 1)] for i in range(num_admins)]

    if solution_limit < 1:
        print('Nem kertel megoldast!')
        
    elif solution_limit == 1:
        # SEGEDVALTOZOK A SZABADNAPOK AZONOSITASARA
        # is_free_day[(a, d)]: IGAZ, HA 'a' ADMIN A 'd' NAPON PIHEN
        is_free_day = {}
        for a in all_admins:
            for d in all_days:
                is_free_day[(a, d)] = model.NewBoolVar(f'is_free_day_a{a}_d{d}')
                # is_free_day[(a,d)] igaz, ha shifts[(a,d,0)] HAMIS ÉS shifts[(a,d,1)] HAMIS
                model.AddBoolAnd([shifts[(a, d, 0)].Not(), shifts[(a, d, 1)].Not()]).OnlyEnforceIf(is_free_day[(a, d)])

        # SEGEDVALTOZOK A DUPLA SZABADNAPOK AZONOSITASARA
        # is_double_free[(n, d)]: IGAZ, HA 'a' ADMIN A 'd' NAPON ÉS A 'd+1' NAPON PIHEN
        is_double_free = {}
        for a in all_admins:
            for d in range(num_days - 1): # AZ UTOLSO NAPRA NEM KELL ELLENORIZNI, MERT NINCS d+1
                is_double_free[(a, d)] = model.NewBoolVar(f'is_double_free_a{a}_d{d}')
                # is_double_free[(a,d)] igaz, ha is_free_day[(a,d)] IGAZ ÉS is_free_day[(a,d+1)] IGAZ
                model.AddBoolAnd([is_free_day[(a, d)], is_free_day[(a, d + 1)]]).OnlyEnforceIf(is_double_free[(a, d)])

        # SEGEDVALTOZOK A TRIPLA SZABADNAPOK AZONOSITASARA
        # is_triple_free[(a, d)]: IGAZ, HA 'a' ADMIN A 'd' NAPON, 'd+1' NAPON ÉS 'd+2' NAPON PIHEN
        # is_triple_free = {}
        # for a in all_admins:
        #     for d in range(num_days - 2): # AZ UTOLSO KET NAPRA NEM KELL ELLENORIZNI
        #         is_triple_free[(a, d)] = model.NewBoolVar(f'is_triple_free_a{a}_d{d}')
        #         # is_triple_free[(a,d)] igaz, ha is_free_day[(a,d)] IGAZ, is_free_day[(a,d+1)] IGAZ ÉS is_free_day[(a,d+2)] IGAZ
        #         model.AddBoolAnd([is_free_day[(a, d)], is_free_day[(a, d + 1)], is_free_day[(a, d + 2)]]).OnlyEnforceIf(is_triple_free[(a, d)])

        # SULYOK A CELFUGGVENYHEZ
        double_free_weight = 8
        triple_free_weight = 6

        # CELFUGGVENY FELEPITESE
        objective_terms = []
        for a in all_admins:
            for d in range(num_days - 1): # DUPLA SZABADNAPOKHOZ
                objective_terms.append(is_double_free[(a, d)] * double_free_weight)
            # for d in range(num_days - 2): # TRIPLA SZABADNAPOKHOZ
            #     objective_terms.append(is_triple_free[(a, d)] * triple_free_weight)

        # MAXIMALIZALJUK A SZABADNAPOK SULYOZOTT OSSZEGET
        model.Maximize(sum(objective_terms))
        
        # SOLVER LETREHOZASA ES MEGOLDAS
        solver = cp_model.CpSolver()
        status = solver.Solve(model)
        
        if status == cp_model.OPTIMAL:
            for a in range(num_admins):
                sd = 0
                sn = 0
                for d in range(num_days):
                    if solver.Value(shifts[(a, d, 0)]) == 1:
                        fin_table[a][d + 1] = 'N'
                        sd += 1
                    elif solver.Value(shifts[(a, d, 1)]) == 1:
                        fin_table[a][d + 1] = 'E'
                        sn += 1
                    else:
                        fin_table[a][d + 1] = ' '
                # fin_table[a].append(sd)
                # fin_table[a].append(sn)
        else:
            print("No optimal solution found !")
            model.Maximize(1)
            
            # SOLVER LETREHOZASA ES MEGOLDAS
            solver = cp_model.CpSolver()
            status = solver.Solve(model)
            
            if status == cp_model.OPTIMAL:
                for a in range(num_admins):
                    sd = 0
                    sn = 0
                    for d in range(num_days):
                        if solver.Value(shifts[(a, d, 0)]) == 1:
                            fin_table[a][d + 1] = 'N'
                            sd += 1
                        elif solver.Value(shifts[(a, d, 1)]) == 1:
                            fin_table[a][d + 1] = 'E'
                            sn += 1
                        else:
                            fin_table[a][d + 1] = ' '
                    # fin_table[a].append(sd)
                    # fin_table[a].append(sn)
            else:
                print('No solution found!')
                return 0

        for i in fin_table:
            print(i)

    else:
        solver = cp_model.CpSolver()
        solution_printer = SolutionPrinter(admins_name, shifts, all_admins, num_days, all_shifts, sat_orig, sun_orig, solution_limit)

        solver.parameters.enumerate_all_solutions = True
        # MEGOLDAS
        status = solver.SearchForAllSolutions(model, solution_printer)

        # print(f"Status = {solver.StatusName(status)}")
        # print(f"Number of solutions found: {solution_printer.solution_count}")



if __name__ == "__main__":
    main()

    ### NAGYOBB ELEMSZAMU FILE MEGTARTASA
    folder = '/home/angyal/dat/prog/python/generalis-admin/cpsat_solver'
    file_tmp = folder + '/admin_schedule_tmp.xlsx'
    file_orig = folder + '/admin_schedule.xlsx'
    
    if os.path.exists(file_orig):
        if os.path.getsize(file_tmp) > os.path.getsize(file_orig):
            os.remove(file_orig)
            os.rename(file_tmp, file_orig)
        else:
            os.remove(file_tmp)
    else:
        os.rename(file_tmp, file_orig)
